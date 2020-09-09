from selenium import webdriver
import openpyxl


"""
this code requiers Firefox 80.0.1 and geckodriver-v0.27.0-win64.
File should be named "config_file" and its spreadsheet "Config".
First row is for Headers: login / passowrd / repository.
Next rows should contain data.

Uses excel file to extract multiple login and passwords.
Logs into GitHub.com selects repo from file and saves screenshot.
Logs out and quits Firefox.
"""

def selerium():
    browser = webdriver.Firefox()
    browser.get("https:github.com")
    sign_in = browser.find_element_by_link_text("Sign in")
    sign_in.click()
    user_name_box = browser.find_element_by_id("login_field")
    user_name_box.send_keys(login)
    password_box = browser.find_element_by_id("password")
    password_box.send_keys(passw)
    password_box.submit()
    browser.implicitly_wait(2)
    repoz = browser.find_element_by_css_selector(
        "div#repos-container > ul.list-style-none > li > div > a > span[title=" + repo + "]")
    repoz.click()
    browser.save_screenshot("screenshot" + str(row) + ".png")
    browser.quit()


workbook = openpyxl.load_workbook("config_file.xlsx")
sheet = workbook["Config"]
for row in range(2, sheet.max_row +1):
    login = sheet.cell(row, 1).value
    passw = sheet.cell(row, 2).value
    repo = sheet.cell(row, 3).value
    if login == None or passw == None or repo == None:
        break
    selerium()

